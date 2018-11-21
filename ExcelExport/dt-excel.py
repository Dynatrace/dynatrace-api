import pycurl
import json
import csv
import certifi
import io
from openpyxl import Workbook
from openpyxl.styles import Alignment,Font

### Setup Variables ###
URL='https://{id}.live.dynatrace.com/api/v1/'
APITOKEN='XXXXXXXXXXXXXXXXXXXXX'
DEST_FILENAME='dt-export.xlsx'





### function to go get the data
def dtApiQuery(endpoint):
	buffer=io.BytesIO()
	c = pycurl.Curl()
	c.setopt(c.URL, URL + endpoint)
	c.setopt(pycurl.CAINFO, certifi.where())
	c.setopt(c.HTTPHEADER, ['Authorization: Api-Token ' + APITOKEN] )
	c.setopt(pycurl.WRITEFUNCTION, buffer.write)
	c.perform()
	print('Status: %d' % c.getinfo(c.RESPONSE_CODE))
	c.close()
	return(buffer.getvalue().decode('UTF-8'))


### Setup workbook
wb = Workbook()
wsHosts = wb.create_sheet("hosts")
wsHostHost = wb.create_sheet("host-host")
wsProcess = wb.create_sheet("processes")
wsProcessProcess = wb.create_sheet("process-process")
wsProcessHost = wb.create_sheet("process-host")
wb.remove(wb.active)



### Get & Process hosts data
hostsIO=dtApiQuery('entity/infrastructure/hosts')
hosts=json.loads(hostsIO)

wsHosts.append( ['hostId','displayName','osType','osVersion','hypervisorType','ipAddress1','ipAddress2','ipAddress3'] )
for host in hosts:
	wsHosts.append( [ host['entityId'],
		host['displayName'],
		host['osType'],
		host['osVersion'],
		host['hypervisorType'] if 'hypervisorType' in host else '',
		host['ipAddresses'][0] if 'ipAddresses' in host else '', 
		host['ipAddresses'][1] if 'ipAddresses' in host and len(host['ipAddresses']) >1 else '',
                host['ipAddresses'][2] if 'ipAddresses' in host and len(host['ipAddresses']) >2 else '' 
	] )

wsHostHost.append( ['fromHostId','toHostId'] )
for fromHost in hosts:
	if 'toRelationships' in fromHost and 'isNetworkClientOfHost' in fromHost['toRelationships']:
		for toHost in fromHost['toRelationships']['isNetworkClientOfHost']:
			wsHostHost.append( [ fromHost['entityId'],
			toHost,
		] )



### Get & Process processes data
processesIO=dtApiQuery('entity/infrastructure/processes')
processes=json.loads(processesIO)

wsProcess.append( ['processId','displayName','softwareType','softwareVersion','port1','port2','port3','port4','port5'] )
for process in processes:
	wsProcess.append( [ process['entityId'],
		process['displayName'],
		process['softwareTechnologies'][0]['type'] if 'softwareTechnologies' in process else '',
		process['softwareTechnologies'][0]['version'] if 'softwareTechnologies' in process else '',
		process['listenPorts'][0] if 'listenPorts' in process else '', 
		process['listenPorts'][1] if 'listenPorts' in process and len(process['listenPorts'])>1 else '', 
		process['listenPorts'][2] if 'listenPorts' in process and len(process['listenPorts'])>2 else '', 
		process['listenPorts'][3] if 'listenPorts' in process and len(process['listenPorts'])>3 else '', 
		process['listenPorts'][4] if 'listenPorts' in process and len(process['listenPorts'])>4 else '' 
	] )

wsProcessProcess.append( ['fromProcessId','toProcessId'] )
for fromProcess in processes:
	if 'toRelationships' in fromProcess and 'isNetworkClientOf' in fromProcess['toRelationships']:
		for toProcess in fromProcess['toRelationships']['isNetworkClientOf']:
			wsProcessProcess.append( [ fromProcess['entityId'],
			toProcess,
		] )

wsProcessHost.append( ['processId','hostId'] )
for process in processes:
	if 'fromRelationships' in process and 'isProcessOf' in process['fromRelationships']:
		for host in process['fromRelationships']['isProcessOf']:
			wsProcessHost.append( [ process['entityId'],
			host,
		] )



### set column widths
for ws in wb.worksheets:
	for column_cells in ws.columns:
    		length = max(len(str(cell.value)) for cell in column_cells)
    		ws.column_dimensions[column_cells[0].column].width = length+1
### Set header format
for ws in wb.worksheets:
	for cell in ws["1:1"]:
		cell.style='Headline 3'


### Generate FW Rule Sheet
wsFWRules = wb.create_sheet("FWRules",0)
wsFWRules.append([ 'Linking Pointers','','','','',
	'Firewall Rule','','','','',
	'Source Extended Info','','','',
	'Destination Extended Info','','','',
	'Filtering'
	])
wsFWRules.merge_cells('A1:E1')
wsFWRules.merge_cells('F1:J1')
wsFWRules.merge_cells('K1:N1')
wsFWRules.merge_cells('O1:R1')
wsFWRules.append([ 'fromProcessId','toProcessId','fromHostId','toHostId','',
	'srcIP','dstIP','proto','port','',
	'srcHostName','srcProcessName','srcProcessType','',
	'dstHostname','dstProcessName','dstProcessType','',
	'isIntraHost?'
	])
for col in ['A','B','C','D','E']:
	wsFWRules.column_dimensions[col].hidden=True
wsFWRules["A1"].style="Accent3"
wsFWRules["F1"].style="Accent1"
wsFWRules["K1"].style="Accent4"
wsFWRules["O1"].style="Accent5"
wsFWRules["S1"].style="Accent2"
for cell in wsFWRules["1:1"]:
	cell.font=Font(bold=True,color='FFFFFF')
	cell.alignment=Alignment(horizontal='center')
for cell in wsFWRules["2:2"]:
	cell.style='Headline 3'
wsFWRules.sheet_properties.tabColor = '0066FF'

i=3
for row in wsProcessProcess.rows:
	wsFWRules.append([
		"='process-process'!A%i" % i,
		"='process-process'!B%i" % i,
		"=VLOOKUP(A%i,'process-host'!$A:$B,2,FALSE)" % i,
		"=VLOOKUP(B%i,'process-host'!$A:$B,2,FALSE)" % i,
		"",
		"=VLOOKUP(C%i,'hosts'!$A:$F,6,FALSE)" % i,
		"=VLOOKUP(D%i,'hosts'!$A:$F,6,FALSE)" % i,
		"TCP",
		"=IF(LEN(VLOOKUP(B%i,'processes'!$A:$E,5,FALSE))=0,\"\",VLOOKUP(B%i,'processes'!$A:$E,5,FALSE))" % (i,i),
		"",
		"=VLOOKUP(C%i,'hosts'!$A:$B,2,FALSE)" % i,
		"=VLOOKUP(A%i,'processes'!$A:$B,2,FALSE)" % i,
		"=VLOOKUP(A%i,'processes'!$A:$C,3,FALSE)" % i,
		"",
		"=VLOOKUP(D%i,'hosts'!$A:$B,2,FALSE)" % i,
		"=VLOOKUP(B%i,'processes'!$A:$B,2,FALSE)" % i,
		"=VLOOKUP(B%i,'processes'!$A:$C,3,FALSE)" % i,
		"",
		"=IF(C%i=D%i,TRUE,FALSE)" % (i,i)
	])
	i += 1
		

wsFWRules.column_dimensions['F'].width = wsHosts.column_dimensions['F'].width
wsFWRules.column_dimensions['G'].width = wsHosts.column_dimensions['F'].width
wsFWRules.column_dimensions['H'].width = 8
wsFWRules.column_dimensions['I'].width = wsProcess.column_dimensions['E'].width
wsFWRules.column_dimensions['J'].width = 5
wsFWRules.column_dimensions['K'].width = wsHosts.column_dimensions['B'].width
wsFWRules.column_dimensions['L'].width = wsProcess.column_dimensions['B'].width
wsFWRules.column_dimensions['M'].width = wsProcess.column_dimensions['C'].width
wsFWRules.column_dimensions['N'].width = 5
wsFWRules.column_dimensions['O'].width = wsHosts.column_dimensions['B'].width
wsFWRules.column_dimensions['P'].width = wsProcess.column_dimensions['B'].width
wsFWRules.column_dimensions['Q'].width = wsProcess.column_dimensions['C'].width
wsFWRules.column_dimensions['R'].width = 5
wsFWRules.column_dimensions['S'].width = 8

wsFWRules.auto_filter.ref="A2:S2"



### Output file
wb.save(filename=DEST_FILENAME)
